from openpyxl import Workbook
from coinex_clinet import CoinexClient

mstats = CoinexClient.get_all_market_stats()['ticker']
print(mstats)
wb = Workbook()
sheet = wb.active
sheet.cell(1, 1).value = 'name'
n = 2
for key in mstats['KAVAUSDT']:
    sheet.cell(1, n).value = key
    n += 1
n = 2
for market, data in mstats.items():
    sheet.cell(n, 1).value = market
    m = 2
    for k in data:
        sheet.cell(n, m).value = data[k]
        m += 1
    n += 1
wb.save('out.xlsx')